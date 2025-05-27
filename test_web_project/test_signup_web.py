import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Định nghĩa đường dẫn file Excel
test_cases_file = r"D:\Tester2\test_web_project\data_report\signup\test_cases_signup.xlsx"
report_file = r"D:\Tester2\test_web_project\data_report\signup\report_signup.xlsx"

# Mở file Excel chứa test case
wb = openpyxl.load_workbook(test_cases_file)
ws = wb.active

# Kiểm tra nếu cột "Actual Result" và "Status" chưa tồn tại thì thêm vào
header_row = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
if "Actual Result" not in header_row:
    ws.cell(row=1, column=ws.max_column + 1, value="Actual Result")
if "Status" not in header_row:
    ws.cell(row=1, column=ws.max_column + 1, value="Status")

# Xác định chỉ mục của các cột
actual_result_col = header_row.index("Actual Result") + 1 if "Actual Result" in header_row else ws.max_column - 1
status_col = header_row.index("Status") + 1 if "Status" in header_row else ws.max_column

# Khởi tạo WebDriver (tự động tải ChromeDriver)
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

# URL trang đăng ký
signup_url = "https://nodejs-ck-x8q8.onrender.com/signup"

# Duyệt qua từng test case trong file Excel
for row in range(2, ws.max_row + 1):
    test_case = ws.cell(row=row, column=1).value  # Test Case ID
    username = ws.cell(row=row, column=3).value   # Username
    email = ws.cell(row=row, column=4).value      # Email
    password = ws.cell(row=row, column=5).value   # Password
    confirm_password = ws.cell(row=row, column=6).value  # Confirm Password
    expected_result = ws.cell(row=row, column=7).value   # Expected Result

    print(f"🔹 Đang kiểm thử: {test_case}")

    # Mở trang đăng ký
    driver.get(signup_url)
    time.sleep(2)

    # Nhập dữ liệu vào form đăng ký
    try:
        name_field = driver.find_element(By.ID, "name")
        email_field = driver.find_element(By.ID, "email")
        password_field = driver.find_element(By.ID, "password")
        confirm_password_field = driver.find_element(By.ID, "confirmpasword")
        signup_button = driver.find_element(By.XPATH, "//button[@type='submit']")

        # Điền thông tin, xử lý trường hợp None
        name_field.send_keys(username if username else "")
        email_field.send_keys(email if email else "")
        password_field.send_keys(password if password else "")
        confirm_password_field.send_keys(confirm_password if confirm_password else "")
        signup_button.click()

        time.sleep(3)  # Chờ trang phản hồi

        # Kiểm tra kết quả
        if driver.current_url != signup_url:
            actual_result = "Sign up successfully"
        else:
            actual_result = "Signup failed"
            try:
                alert_box = driver.find_element(By.CLASS_NAME, "alert")
                actual_result = alert_box.text
            except:
                pass  # Giữ nguyên "Signup failed" nếu không tìm thấy alert

        # So sánh kết quả
        status = "Pass" if actual_result.lower() == expected_result.lower() else "Fail"

        # Ghi kết quả vào file Excel
        ws.cell(row=row, column=actual_result_col, value=actual_result)
        ws.cell(row=row, column=status_col, value=status)

        print(f"✅ Expected: {expected_result} | Actual: {actual_result} | Status: {status}")

    except Exception as e:
        print(f"❌ Lỗi khi kiểm thử: {test_case} - {str(e)}")
        ws.cell(row=row, column=actual_result_col, value="Test Failed")
        ws.cell(row=row, column=status_col, value="Fail")

# Lưu file báo cáo
wb.save(report_file)
wb.close()
driver.quit()

print(f"🎯 Hoàn thành kiểm thử! Báo cáo được lưu tại: {report_file}")