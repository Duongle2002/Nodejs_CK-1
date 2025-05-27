import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# Đọc dữ liệu từ file Excel đầu vào
input_file = r"D:\nam3\hk2\KTPM2\test_web_project\test_web_project\data_report\search\test_cases_search.xlsx"
output_file = r"D:\nam3\hk2\KTPM2\test_web_project\test_web_project\data_report\search\report_search.xlsx"

# Mở file Excel chứa danh sách từ khóa tìm kiếm
wb_input = openpyxl.load_workbook(input_file)
ws_input = wb_input.active

# Tạo file Excel mới để ghi kết quả
wb_output = openpyxl.Workbook()
ws_output = wb_output.active
# Tiêu đề với cột Test Case ID và Test Case ở đầu
ws_output.append(["Test Case ID", "Test Case", "Search Query", "Expected Result", "Actual Result", "Test Status", "Product Count", "Product Names"])

# Khởi tạo WebDriver
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")  # Mở toàn màn hình
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Mở trang web
driver.get("https://nodejs-ck-x8q8.onrender.com/product")  # ⚠️ Thay bằng URL trang web của bạn

# Lặp qua từng dòng trong file Excel (bắt đầu từ hàng thứ 2)
for row in ws_input.iter_rows(min_row=2, values_only=True):
    test_case_id = row[0]  # Cột chứa Test Case ID (TC_SR_001, TC_SR_002, ...)
    test_case = row[1]     # Cột chứa Test Case (mô tả test case)
    search_query = row[2]  # Cột chứa từ khóa tìm kiếm
    expected_result = row[3].strip()  # Cột chứa kết quả mong đợi (loại bỏ khoảng trắng)

    try:
        # Tìm ô tìm kiếm
        search_input = driver.find_element(By.NAME, "search")
        search_input.clear()  # Xóa nội dung cũ
        if search_query is not None:  # Kiểm tra nếu search_query không phải None
            search_input.send_keys(search_query)
        search_input.send_keys(Keys.RETURN)  # Nhấn Enter

        # Chờ trang cập nhật kết quả
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.ID, "productList")))
        time.sleep(3)  # Thêm thời gian chờ 3 giây

        # Kiểm tra số lượng sản phẩm hiển thị
        product_list = driver.find_element(By.ID, "productList")  
        product_elements = product_list.find_elements(By.CLASS_NAME, "col-md-4")  # Danh sách sản phẩm

        product_count = len(product_elements)  # Đếm số sản phẩm
        product_names = []

        # Lấy tên sản phẩm
        for product in product_elements:
            try:
                name_element = product.find_element(By.TAG_NAME, "h2")
                product_names.append(name_element.text.strip())  # Lấy tên sản phẩm
            except:
                continue  # Bỏ qua nếu không tìm thấy tên sản phẩm

        # Tạo chuỗi kết quả giống Expected Result
        if product_count > 0:
            actual_result = f"{product_count} | {', '.join(product_names)}"
        else:
            actual_result = "0 | No results found"

        # So sánh với Expected Result
        test_status = "PASS" if actual_result.strip() == expected_result.strip() else "FAIL"

        # Ghi kết quả vào file Excel, bao gồm Test Case ID và Test Case
        ws_output.append([test_case_id, test_case, search_query, expected_result, actual_result, test_status, product_count, ', '.join(product_names)])

    except Exception as e:
        print(f"❌ Lỗi khi kiểm thử với Test Case ID '{test_case_id}': {e}")
        # Ghi lỗi vào file Excel, bao gồm Test Case ID và Test Case
        ws_output.append([test_case_id, test_case, search_query, expected_result, "Error", "FAIL", 0, ""])

# Lưu file Excel kết quả
wb_output.save(output_file)
print(f"✅ Kiểm thử hoàn tất! Báo cáo lưu tại: {output_file}")

# Đóng trình duyệt
driver.quit()