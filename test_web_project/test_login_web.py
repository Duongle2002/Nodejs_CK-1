import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Định nghĩa đường dẫn file Excel
test_cases_file = r"D:\nam3\hk2\KTPM2\test_web_project\data_report\login\Login_Test_Cases.xlsx"
report_file = r"D:\Tester2\test_web_project\data_report\login\Report_Login.xlsx"

# Mở file Excel chứa test case
wb = openpyxl.load_workbook(test_cases_file)
ws = wb.active

# Thêm cột mới "Actual Result" và "Status" nếu chưa có
headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
if "Actual Result" not in headers:
    ws.cell(row=1, column=ws.max_column + 1, value="Actual Result")
if "Status" not in headers:
    ws.cell(row=1, column=ws.max_column + 2, value="Status")

# Xác định vị trí các cột
col_username = headers.index("Username") + 1
col_password = headers.index("Password") + 1
col_expected = headers.index("Expected Result") + 1
col_actual = ws.max_column - 1
col_status = ws.max_column

# Khởi tạo WebDriver
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

# URL trang đăng nhập
login_url = "https://nodejs-ck-x8q8.onrender.com/login"

# Duyệt qua từng test case
for row in range(2, ws.max_row + 1):
    test_case = ws.cell(row=row, column=1).value
    username = ws.cell(row=row, column=col_username).value or ""
    password = ws.cell(row=row, column=col_password).value or ""
    expected_result = ws.cell(row=row, column=col_expected).value

    print(f"🔍 Đang kiểm thử: {test_case}")

    # Mở trang đăng nhập
    driver.get(login_url)
    time.sleep(2)

    try:
        # Tìm và nhập thông tin đăng nhập
        email_field = driver.find_element(By.ID, "email")
        password_field = driver.find_element(By.ID, "password")
        login_button = driver.find_element(By.XPATH, "//button[@type='submit']")

        email_field.send_keys(username)
        password_field.send_keys(password)
        login_button.click()

        time.sleep(3)  # Chờ phản hồi từ server

        # Kiểm tra nếu URL thay đổi (đăng nhập thành công)
        current_url = driver.current_url
        if current_url != login_url:
            actual_result = "Logged in successfully"
        else:
            # Nếu URL không đổi, tìm thông báo lỗi
            try:
                alert_box = driver.find_element(By.CLASS_NAME, "alert")  # Điều chỉnh theo UI của bạn
                actual_result = "Login failed"
            except:
                actual_result = "Login failed"

        # So sánh kết quả thực tế với mong đợi
        status = "Pass" if actual_result == expected_result else "Fail"

        # Ghi kết quả vào file Excel
        ws.cell(row=row, column=col_actual, value=actual_result)
        ws.cell(row=row, column=col_status, value=status)

        print(f"✅ Expected: {expected_result} | Actual: {actual_result} | Status: {status}")

    except Exception as e:
        print(f"❌ Lỗi khi kiểm thử: {e}")
        ws.cell(row=row, column=col_actual, value="Test Failed")
        ws.cell(row=row, column=col_status, value="Fail")

# Lưu file báo cáo
# Lưu file báo cáo
import os
os.makedirs(os.path.dirname(report_file), exist_ok=True)
wb.save(report_file)
wb.close()
driver.quit()

print(f"📊 Đã hoàn thành kiểm thử! Báo cáo được lưu tại: {report_file}")

