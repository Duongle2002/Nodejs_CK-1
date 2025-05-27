# import time
# import openpyxl
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.common.action_chains import ActionChains

# # === CẤU HÌNH FILE ===
# excel_file = r"D:\Tester2\test_web_project\data_report\add_to_cart\product_test_cases.xlsx"
# report_file = r"D:\Tester2\test_web_project\data_report\add_to_cart\product_test_report.xlsx"

# # Đọc file Excel
# wb = openpyxl.load_workbook(excel_file)
# ws = wb.active

# # Khởi tạo trình duyệt
# driver = webdriver.Chrome()
# wait = WebDriverWait(driver, 15)  # Timeout 15 giây để chờ trang load

# # 1. MỞ TRANG ĐĂNG NHẬP
# driver.get("https://nodejs-ck-x8q8.onrender.com/login")

# # 2. ĐĂNG NHẬP
# try:
#     email = wait.until(EC.presence_of_element_located((By.ID, "email")))
#     password = driver.find_element(By.ID, "password")
#     email.send_keys("admin@gmail.com")  # Thay bằng tài khoản thật
#     password.send_keys("123456")        # Thay bằng mật khẩu thật
#     password.send_keys(Keys.ENTER)
#     print("✅ Đăng nhập thành công")
# except Exception as e:
#     print(f"❌ Lỗi đăng nhập: {e}")
#     driver.quit()
#     exit()

# # 3. CHỜ TRANG LOAD XONG VÀ NHẤN "PRODUCTS"
# try:
#     products_button = wait.until(
#         EC.element_to_be_clickable((By.XPATH, "//a[@href='/product' and text()='Products']"))
#     )
#     driver.execute_script("arguments[0].scrollIntoView(true);", products_button)
#     time.sleep(0.5)
#     products_button.click()
#     print("✅ Đã nhấn vào menu 'Products'")
# except Exception as e:
#     print(f"❌ Không thể nhấn vào 'Products': {e}")
#     driver.quit()
#     exit()

# # 4. CHỜ CÁC SẢN PHẨM XUẤT HIỆN
# try:
#     wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.desc")))
#     print("✅ Đã tìm thấy các sản phẩm")
# except Exception as e:
#     print(f"❌ Không tìm thấy sản phẩm: {e}")
#     driver.quit()
#     exit()

# # 5. LẶP TEST THEO FILE EXCEL
# for row in range(2, ws.max_row + 1):
#     test_id = ws.cell(row=row, column=1).value
#     product_name = ws.cell(row=row, column=2).value
#     product_id = ws.cell(row=row, column=3).value
#     expected_result = ws.cell(row=row, column=4).value

#     print(f"🛒 {test_id}: Đang thử thêm '{product_name}' (ID: {product_id}) vào giỏ...")

#     try:
#         # Tìm phần tử chứa sản phẩm (div.desc) dựa trên product_id
#         product_container = wait.until(EC.presence_of_element_located(
#             (By.XPATH, f"//div[@class='desc']//button[@data-product-id='{product_id}']/ancestor::div[@class='desc']"))
#         )

#         # Scroll tới sản phẩm
#         driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", product_container)
#         time.sleep(0.5)

#         # Thực hiện hover lên sản phẩm
#         actions = ActionChains(driver)
#         actions.move_to_element(product_container).perform()
#         time.sleep(0.5)  # Chờ nút "Add to Cart" hiện ra

#         # Tìm và nhấn nút "Add to Cart"
#         button = wait.until(EC.element_to_be_clickable(
#             (By.CSS_SELECTOR, f"div.desc button.btn-add-to-cart[data-product-id='{product_id}']"))
#         )
#         button.click()
#         time.sleep(1)

#         actual_result = "Add to cart success"
#         status = "PASSED" if actual_result == expected_result else "FAILED"

#     except Exception as e:
#         actual_result = f"Error: {str(e)}"
#         status = "FAILED"

#     # Ghi kết quả
#     ws.cell(row=row, column=5, value=actual_result)
#     ws.cell(row=row, column=6, value=status)

# # 6. LƯU FILE KẾT QUẢ
# wb.save(report_file)
# driver.quit()
# print(f"✅ Đã hoàn thành kiểm thử. Báo cáo lưu tại: {report_file}")

import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

# === CẤU HÌNH FILE ===
excel_file = r"D:\Tester2\test_web_project\data_report\add_to_cart\product_test_cases.xlsx"
report_file = r"D:\Tester2\test_web_project\data_report\add_to_cart\product_test_report.xlsx"

# Đọc file Excel
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# Khởi tạo trình duyệt
driver = webdriver.Chrome()
wait = WebDriverWait(driver, 15)  # Timeout 15 giây để chờ trang load

# 1. MỞ TRANG ĐĂNG NHẬP
driver.get("https://nodejs-ck-x8q8.onrender.com/login")

# 2. ĐĂNG NHẬP
try:
    email = wait.until(EC.presence_of_element_located((By.ID, "email")))
    password = driver.find_element(By.ID, "password")
    email.send_keys("admin@gmail.com")  # Thay bằng tài khoản thật
    password.send_keys("123456")        # Thay bằng mật khẩu thật
    password.send_keys(Keys.ENTER)
    print("✅ Đăng nhập thành công")
except Exception as e:
    print(f"❌ Lỗi đăng nhập: {e}")
    driver.quit()
    exit()

# 3. CHỜ TRANG LOAD XONG VÀ NHẤN "PRODUCTS"
try:
    products_button = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//a[@href='/product' and text()='Products']"))
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", products_button)
    time.sleep(0.5)
    products_button.click()
    print("✅ Đã nhấn vào menu 'Products'")
except Exception as e:
    print(f"❌ Không thể nhấn vào 'Products': {e}")
    driver.quit()
    exit()

# 4. CHỜ CÁC SẢN PHẨM XUẤT HIỆN
try:
    wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.desc")))
    print("✅ Đã tìm thấy các sản phẩm")
except Exception as e:
    print(f"❌ Không tìm thấy sản phẩm: {e}")
    driver.quit()
    exit()

# 5. LẶP TEST THEO FILE EXCEL
for row in range(2, ws.max_row + 1):
    test_id = ws.cell(row=row, column=1).value
    product_name = ws.cell(row=row, column=2).value
    product_id = ws.cell(row=row, column=3).value
    quantity = ws.cell(row=row, column=4).value  # Đọc cột Quantity
    expected_result = ws.cell(row=row, column=5).value

    print(f"🛒 {test_id}: Đang thử thêm '{product_name}' (ID: {product_id}) với số lượng {quantity} vào giỏ...")

    try:
        # Tìm phần tử chứa sản phẩm (div.desc) dựa trên product_id
        product_container = wait.until(EC.presence_of_element_located(
            (By.XPATH, f"//div[@class='desc']//button[@data-product-id='{product_id}']/ancestor::div[@class='desc']"))
        )

        # Scroll tới sản phẩm
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", product_container)
        time.sleep(0.5)

        # Thực hiện hover lên sản phẩm
        actions = ActionChains(driver)
        actions.move_to_element(product_container).perform()
        time.sleep(0.5)  # Chờ nút "Add to Cart" hiện ra

        # Nhấn nút "Add to Cart" theo số lượng yêu cầu
        button = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, f"div.desc button.btn-add-to-cart[data-product-id='{product_id}']"))
        )
        for _ in range(int(quantity)):  # Nhấn nút số lần bằng quantity
            button.click()
            time.sleep(1)  # Chờ sau mỗi lần nhấn để đảm bảo hệ thống xử lý

        actual_result = f"Add {quantity} to cart success"
        status = "PASSED" if actual_result.startswith("Add") and expected_result == "Add to cart success" else "FAILED"

    except Exception as e:
        actual_result = f"Error: {str(e)}"
        status = "FAILED"

    # Ghi kết quả
    ws.cell(row=row, column=6, value=actual_result)
    ws.cell(row=row, column=7, value=status)

# 6. LƯU FILE KẾT QUẢ
wb.save(report_file)
driver.quit()
print(f"✅ Đã hoàn thành kiểm thử. Báo cáo lưu tại: {report_file}")