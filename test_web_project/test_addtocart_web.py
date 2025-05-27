import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Đọc dữ liệu từ file Excel đầu vào
input_file = r"D:\Tester2\test_web_project\data_report\add_to_cart\add_to_cart_test_cases.xlsx"
output_file = r"D:\Tester2\test_web_project\data_report\add_to_cart\report_cart.xlsx"

# Mở file Excel chứa danh sách test case
wb_input = openpyxl.load_workbook(input_file)
ws_input = wb_input.active

# Tạo file Excel mới để ghi kết quả
wb_output = openpyxl.Workbook()
ws_output = wb_output.active
ws_output.append(["Email", "Password", "Product Name", "Expected Result", "Actual Result", "Test Status"])

# Khởi tạo WebDriver
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Lặp qua từng dòng trong file Excel
for row in ws_input.iter_rows(min_row=2, values_only=True):
    email, password, product_name, expected_result = row
    
    try:
        # Mở trang đăng nhập
        driver.get("https://nodejs-ck-x8q8.onrender.com/login")
        time.sleep(2)

        # Nhập email và password
        driver.find_element(By.ID, "email").send_keys(email)
        driver.find_element(By.ID, "password").send_keys(password)
        driver.find_element(By.XPATH, "//button[text()='Sign in']").click()

        # Chờ đăng nhập thành công
        time.sleep(3)

        # Mở trang sản phẩm
        driver.get("https://nodejs-ck-x8q8.onrender.com/product")
        time.sleep(5)

        # Lấy danh sách tất cả sản phẩm
        products = driver.find_elements(By.CLASS_NAME, "product-item")
        print("\n🔍 Danh sách sản phẩm hiển thị:")
        for p in products:
            print("-", p.find_element(By.TAG_NAME, "h2").text.strip())

        # Tìm sản phẩm cần mua
        product_found = False
        for product in products:
            product_name_element = product.find_element(By.TAG_NAME, "h2")
            if product_name_element.text.strip() == product_name.strip():
                product_found = True
                print(f"✅ Đã tìm thấy sản phẩm: {product_name}")

                # Click vào nút Add to Cart chính xác
                add_to_cart_button = product.find_element(By.CLASS_NAME, "btn-add-to-cart")
                add_to_cart_button.click()
                print(f"🛒 Đã bấm 'Add to Cart' cho sản phẩm {product_name}")
                break

        if not product_found:
            actual_result = "Product not found"
            ws_output.append([email, password, product_name, expected_result, actual_result, "FAIL"])
            continue

        # Chờ sản phẩm thêm vào giỏ hàng
        time.sleep(3)

        # Mở trang giỏ hàng
        driver.get("https://nodejs-ck-x8q8.onrender.com/cart")
        time.sleep(3)

        # Kiểm tra sản phẩm trong giỏ hàng
        cart_items = driver.find_elements(By.TAG_NAME, "h2")
        cart_product_names = [item.text.strip() for item in cart_items]
        print("\n🛒 Danh sách sản phẩm trong giỏ hàng:", cart_product_names)

        if product_name in cart_product_names:
            actual_result = "Product added successfully"
        else:
            actual_result = "Product not in cart"

        # So sánh với Expected Result
        test_status = "PASS" if actual_result == expected_result else "FAIL"

        # Ghi kết quả vào file Excel
        ws_output.append([email, password, product_name, expected_result, actual_result, test_status])

    except Exception as e:
        ws_output.append([email, password, product_name, expected_result, "Error", "FAIL"])
        print(f"❌ Lỗi khi kiểm thử với sản phẩm '{product_name}': {e}")

# Lưu file Excel kết quả
wb_output.save(output_file)
print(f"\n✅ Kiểm thử hoàn tất! Báo cáo lưu tại: {output_file}")

# Đóng trình duyệt
driver.quit() 